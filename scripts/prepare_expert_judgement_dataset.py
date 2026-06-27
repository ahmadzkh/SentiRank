#!/usr/bin/env python3
"""
MS-17A — Expert Judgement Spreadsheet Intake & Mapping.

Reads the augmented Google Form spreadsheet, extracts respondent profiles,
pairwise comparison matrices, computes AHP consistency, and writes
processed CSV/JSON outputs for downstream AHP/Fuzzy AHP calculation (MS-17B).

Usage:
    python scripts/prepare_expert_judgement_dataset.py
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent

CRITERIA = {
    "C1": "Features, Content & Audio Experience",
    "C2": "Ads Experience",
    "C3": "Subscription & Pricing",
    "C4": "Account/Login",
    "C5": "App Reliability & Usability",
}

# P01–P10 mapping: (comparison_id, criterion_a, criterion_b, column_name)
PAIRWISE_QUESTIONS = [
    ("P01", "C1", "C2", "P01_C1_vs_C2"),
    ("P02", "C1", "C3", "P02_C1_vs_C3"),
    ("P03", "C1", "C4", "P03_C1_vs_C4"),
    ("P04", "C1", "C5", "P04_C1_vs_C5"),
    ("P05", "C2", "C3", "P05_C2_vs_C3"),
    ("P06", "C2", "C4", "P06_C2_vs_C4"),
    ("P07", "C2", "C5", "P07_C2_vs_C5"),
    ("P08", "C3", "C4", "P08_C3_vs_C4"),
    ("P09", "C3", "C5", "P09_C3_vs_C5"),
    ("P10", "C4", "C5", "P10_C4_vs_C5"),
]

# Pairwise display labels extracted from Form Responses 2 headers
PAIRWISE_DISPLAY_LABELS = {
    "P01": "C1 vs C2",
    "P02": "C1 vs C3",
    "P03": "C1 vs C4",
    "P04": "C1 vs C5",
    "P05": "C2 vs C3",
    "P06": "C2 vs C4",
    "P07": "C2 vs C5",
    "P08": "C3 vs C4",
    "P09": "C3 vs C5",
    "P10": "C4 vs C5",
}

# AHP Saaty's Random Index
RANDOM_INDEX_BY_SIZE = {1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12,
                        6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}

CR_THRESHOLD = 0.10

# Scale mapping from text → numeric value
SCALE_TEXT_TO_NUM = {
    "1": 1.0,
    "3": 3.0,
    "5": 5.0,
    "7": 7.0,
    "9": 9.0,
}

# Regex patterns for parsing Google Form text answers
RE_C1_MORE = re.compile(
    r"C1\s+(?:mutlak\s+)?(?:sangat\s+)?lebih penting (?:daripada )?C2\s*[—–-]+\s*(\d+)"
)
RE_C2_MORE = re.compile(
    r"C2\s+(?:mutlak\s+)?(?:sangat\s+)?lebih penting (?:daripada )?C1\s*[—–-]+\s*(\d+)"
)
# Generic: "C{x} ... lebih penting daripada C{y} — {value}"
RE_GENERIC = re.compile(
    r"C(\d)\s.*?lebih penting.*?C(\d)\s*[—–-]+\s*(\d+)"
)
RE_EQUAL = re.compile(
    r"(?:Keduanya\s+)?sama penting(?:nya)?(?:\s*[—–-]+\s*(\d+))?"
)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class RespondentProfile:
    """Cleaned, anonymized respondent profile."""
    respondent_id: str
    original_code: str
    source_type: str              # "actual" or "synthetic"
    role_category: str
    expertise: str
    education: str
    experience: str
    spotify_status: str
    spotify_frequency: str
    criteria_adequacy: str
    top_criterion: str
    top_reason: str


@dataclass
class PairwiseValues:
    """Pairwise comparison matrix values for one respondent."""
    respondent_id: str
    original_code: str
    source_type: str
    values: dict[str, float]      # {question_id: value}  e.g. {"P01": 9.0}
    matrix: list[list[float]]     # 5x5 reciprocal matrix


@dataclass
class ConsistencyResult:
    """Consistency check result."""
    cr: float
    lambda_max: float
    ci: float
    is_consistent: bool


@dataclass
class RespondentJudgement:
    """Full judgement data for one respondent."""
    profile: RespondentProfile
    pairwise: PairwiseValues
    consistency: ConsistencyResult
    raw_text_answers: dict[str, str] = field(default_factory=dict)
    parse_errors: list[str] = field(default_factory=list)
    validation_notes: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# AHP consistency calculation (mirrors decision-service logic)
# ---------------------------------------------------------------------------


def calculate_ahp_weights(matrix: list[list[float]]) -> list[float]:
    """Geometric mean method."""
    n = len(matrix)
    geometric_means = [math.prod(row) ** (1.0 / n) for row in matrix]
    total = sum(geometric_means)
    if total <= 0:
        raise ValueError("Geometric mean total must be positive.")
    return [v / total for v in geometric_means]


def calculate_consistency(matrix: list[list[float]], weights: list[float]) -> ConsistencyResult:
    """Compute lambda_max, CI, CR for a reciprocal matrix."""
    n = len(matrix)
    if n <= 2:
        return ConsistencyResult(cr=0.0, lambda_max=float(n), ci=0.0, is_consistent=True)

    weighted_sum = [sum(row[i] * weights[i] for i in range(n)) for row in matrix]
    lambda_max = sum(weighted_sum[i] / weights[i] for i in range(n)) / n
    ci = (lambda_max - n) / (n - 1)
    ri = RANDOM_INDEX_BY_SIZE.get(n, 0.0)
    cr = 0.0 if ri == 0 else ci / ri
    return ConsistencyResult(
        cr=cr,
        lambda_max=lambda_max,
        ci=ci,
        is_consistent=cr <= CR_THRESHOLD,
    )


def build_pairwise_matrix(values: dict[str, float], criteria_order: list[str]) -> list[list[float]]:
    """Build 5x5 reciprocal matrix from 10 pairwise values.

    values keys are "P01".."P10", each value is criterion_a / criterion_b
    (a_over_b). The function places reciprocal entries.
    """
    n = len(criteria_order)
    id_to_index = {cid: i for i, cid in enumerate(criteria_order)}
    matrix = [[1.0 if r == c else 0.0 for c in range(n)] for r in range(n)]

    for (qid, cid_a, cid_b, _), display in zip(PAIRWISE_QUESTIONS, PAIRWISE_QUESTIONS):
        val = values.get(qid)
        if val is None or val <= 0:
            continue
        i = id_to_index[cid_a]
        j = id_to_index[cid_b]
        matrix[i][j] = val
        matrix[j][i] = 1.0 / val

    return matrix


# ---------------------------------------------------------------------------
# Spreadsheet parsing
# ---------------------------------------------------------------------------


def parse_text_answer(text: str) -> tuple[float | None, list[str]]:
    """Parse a Google Form pairwise answer text into a numeric value.

    Returns (a_over_b_value, errors_list).
    For 'C1 lebih penting daripada C2 — 5': returns 5.0
    For 'C2 lebih penting daripada C1 — 5': returns 1/5 = 0.2
    For 'Keduanya sama penting — 1': returns 1.0
    """
    errors = []
    if not text or not isinstance(text, str):
        return None, ["Empty answer"]

    text = text.strip()

    # Equal importance
    m = RE_EQUAL.match(text)
    if m:
        return 1.0, []

    # Generic pattern: C{x} ... lebih penting ... C{y} — {value}
    m = RE_GENERIC.match(text)
    if m:
        c_a = int(m.group(1))
        c_b = int(m.group(2))
        raw_val = m.group(3)
        num = SCALE_TEXT_TO_NUM.get(raw_val)
        if num is None:
            return None, [f"Unknown scale value '{raw_val}' in '{text}'"]
        # If C_a is more important than C_b, return num (a_over_b)
        # If C_b is more important than C_a, return 1/num
        # We need to figure out direction
        # RE_GENERIC matches "C{x} ... lebih penting daripada C{y} — {v}"
        # So c_a is the more important one, c_b is the less important
        # But which question is this? We don't know the question from text alone.
        # We'll do a smarter parse below
        question_text = text  # we need context
        _ = question_text  # placeholder

    # Use directional patterns
    # Pattern: "C1 ... lebih penting daripada C2 — {value}" → C1 > C2 → a_over_b = value
    m = re.match(
        r"C(\d)\s+(?:mutlak\s+)?(?:sangat\s+)?lebih penting (?:daripada )?C(\d)\s*[—–-]+\s*(\d+)",
        text,
    )
    if m:
        c_a = int(m.group(1))
        c_b = int(m.group(2))
        raw_val = m.group(3)
        num = SCALE_TEXT_TO_NUM.get(raw_val)
        if num is None:
            return None, [f"Unknown scale value '{raw_val}' in '{text}'"]
        # c_a is more important than c_b → a_over_b = num
        # But this is relative to the specific question's order.
        # For the generic parse, we return a_over_b as-is:
        # the consumer knows which question this is.
        return num, []

    # Pattern: text just has number, assume equal
    m2 = re.match(r"^\s*(\d+)\s*$", text)
    if m2:
        num = SCALE_TEXT_TO_NUM.get(m2.group(1))
        if num:
            return num, []
        return None, [f"Numeric value '{m2.group(1)}' not in allowed scale"]

    return None, [f"Unparseable answer: '{text[:80]} ...'"]


def parse_criteria_pair(
    question_id: str, text_answer: str
) -> tuple[float | None, list[str]]:
    """Parse a specific pairwise question text into the a_over_b value as defined
    by PAIRWISE_QUESTIONS order.

    Returns (a_over_b_value, errors).
    """
    errors = []
    if not text_answer or not isinstance(text_answer, str):
        return None, ["Empty answer"]

    text_answer = text_answer.strip()

    # Equal → 1.0
    if "sama penting" in text_answer:
        return 1.0, []

    # Determine which criteria this question compares
    # For P01: C1 vs C2 → we need C1_over_C2
    # text: "C1 ... lebih penting ... C2 — 5" → C1 more important → 5
    # text: "C2 ... lebih penting ... C1 — 5" → C2 more important → 1/5

    qinfo = None
    for q in PAIRWISE_QUESTIONS:
        if q[0] == question_id:
            qinfo = q
            break
    if qinfo is None:
        return None, [f"Unknown question {question_id}"]

    cid_a, cid_b = qinfo[1], qinfo[2]

    # Extract the numeric value from text
    m = re.search(r"[—–-]+\s*(\d+)", text_answer)
    if not m:
        return None, [f"No numeric value found in '{text_answer}'"]
    raw_val = m.group(1)
    num = SCALE_TEXT_TO_NUM.get(raw_val)
    if num is None:
        return None, [f"Scale value '{raw_val}' not in 1/3/5/7/9"]

    a_num = int(cid_a[1])  # 1..5
    b_num = int(cid_b[1])

    # Find which criterion is mentioned first in the text
    # "C{x} ... lebih penting daripada C{y}" means C{x} is more important
    first_c_match = re.match(r"C(\d)", text_answer)
    if not first_c_match:
        return None, [f"Cannot determine direction from '{text_answer}'"]
    first_c = int(first_c_match.group(1))

    # If the first criterion mentioned equals cid_a, then a_over_b = num
    # If the first criterion equals cid_b, then b_over_a = num → a_over_b = 1/num
    if first_c == a_num:
        return num, []
    elif first_c == b_num:
        return 1.0 / num, []
    else:
        # Unusual: maybe rephrase mentions a different criterion
        return None, [f"Direction mismatch: first criterion in text is C{first_c}, "
                      f"expected C{a_num} or C{b_num}"]


# ---------------------------------------------------------------------------
# Spreadsheet loader
# ---------------------------------------------------------------------------


def load_spreadsheet(path: str | Path) -> openpyxl.Workbook:
    """Load and validate the augmented spreadsheet."""
    path = Path(path)
    if not path.exists():
        print(f"ERROR: Spreadsheet not found: {path}", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(path)
    required = {"Form Responses 2", "Pairwise_Values", "AHP_Validation"}
    missing = required - set(wb.sheetnames)
    if missing:
        print(f"ERROR: Missing sheets: {missing}", file=sys.stderr)
        sys.exit(1)
    return wb


def load_form_responses(ws) -> dict[int, dict]:
    """Load Form Responses 2 sheet into dict keyed by row number (1-indexed)."""
    rows: dict[int, dict] = {}
    headers = [cell.value for cell in ws[1]]
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c, header in enumerate(headers):
            cell = ws[r][c] if c < len(ws[r]) else None
            row_data[header] = cell.value if cell is not None else None
        # Skip fully empty rows
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_data.values()):
            continue
        rows[r] = row_data
    return rows


def load_pairwise_values(ws) -> dict[int, dict]:
    """Load Pairwise_Values sheet keyed by row_in_form (the Google Form row)."""
    rows: dict[int, dict] = {}
    headers = [cell.value for cell in ws[1]]
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c, header in enumerate(headers):
            cell = ws[r][c] if c < len(ws[r]) else None
            row_data[header] = cell.value if cell is not None else None
        row_num = row_data.get("row_in_form")
        if row_num is not None:
            rows[int(row_num)] = row_data
    return rows


def load_ahp_validation(ws) -> dict[int, dict]:
    """Load AHP_Validation sheet keyed by row_in_form."""
    rows: dict[int, dict] = {}
    headers = [cell.value for cell in ws[1]]
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c, header in enumerate(headers):
            cell = ws[r][c] if c < len(ws[r]) else None
            row_data[header] = cell.value if cell is not None else None
        row_num = row_data.get("row_in_form")
        if row_num is not None:
            rows[int(row_num)] = row_data
    return rows


# ---------------------------------------------------------------------------
# Respondent processing pipeline
# ---------------------------------------------------------------------------


def process_respondent(
    form_row_num: int,
    form_row: dict,
    pair_row: dict | None,
    valid_row: dict | None,
    index: int,
) -> RespondentJudgement:
    """Process one respondent through the full intake pipeline."""
    errors: list[str] = []
    notes: list[str] = []

    # --- 1. Determine source type ---
    source_type_raw = (pair_row or valid_row or {}).get("source_type", "unknown")
    if source_type_raw == "actual_uploaded_response":
        source_type = "actual"
    elif source_type_raw == "synthetic_simulation":
        source_type = "synthetic"
    else:
        source_type = "unknown"

    # --- 2. Get original code ---
    original_code = (pair_row or valid_row or {}).get("respondent_code", f"ROW{form_row_num}")

    # --- 3. Anonymized ID ---
    respondent_id = f"EJ{index:03d}"

    # --- 4. Profile ---
    name = (form_row.get("Nama Responden  ") or "").strip()
    profile = RespondentProfile(
        respondent_id=respondent_id,
        original_code=original_code,
        source_type=source_type,
        role_category=(form_row.get("Peran atau Pekerjaan Saat Ini  ") or "").strip(),
        expertise=(form_row.get("Bidang Keahlian atau Pengalaman yang Paling Relevan  ") or "").strip(),
        education=(form_row.get("Pendidikan Terakhir  ") or "").strip(),
        experience=(form_row.get("Lama pengalaman pada bidang yang relevan  ") or "").strip(),
        spotify_status=(form_row.get("Status penggunaan Spotify  ") or "").strip(),
        spotify_frequency=(form_row.get("Frekuensi penggunaan Spotify  ") or "").strip(),
        criteria_adequacy=(form_row.get(
            "Menurut Anda, apakah lima kriteria di atas sudah mewakili masalah utama "
            "yang sering muncul pada aplikasi Spotify? "
        ) or "").strip(),
        top_criterion=(form_row.get(
            "Menurut Anda, kriteria mana yang paling penting untuk diprioritaskan "
            "dalam perbaikan aplikasi Spotify?  "
        ) or "").strip(),
        top_reason=(form_row.get(
            "Berikan alasan singkat atas pilihan prioritas Anda.  "
        ) or "").strip(),
    )

    # --- 5. Pairwise values from Pairwise_Values sheet (primary) ---
    pairwise_values: dict[str, float] = {}
    raw_text_answers: dict[str, str] = {}
    parse_issues: list[str] = []

    if pair_row:
        for qid, _, _, col_name in PAIRWISE_QUESTIONS:
            val = pair_row.get(col_name)
            if val is not None:
                pairwise_values[qid] = float(val)

    # --- 6. Parse text from Form Responses 2 as cross-check ---
    col_prefix = {
        "P01": 15, "P02": 16, "P03": 17, "P04": 18,
        "P05": 19, "P06": 20, "P07": 21, "P08": 22,
        "P09": 23, "P10": 24,
    }

    # Get form headers list to map by index
    form_headers = list(form_row.keys())
    for qid, col_idx in col_prefix.items():
        if col_idx < len(form_headers):
            text_val = form_row.get(form_headers[col_idx])
            if text_val is not None:
                raw_text_answers[qid] = str(text_val).strip()
                # Parse as cross-check
                parsed, parse_errs = parse_criteria_pair(qid, str(text_val))
                if parsed is not None and qid in pairwise_values:
                    # Cross-check: should be similar within 1e-6
                    actual = pairwise_values[qid]
                    if abs(parsed - actual) > 1e-6:
                        parse_issues.append(
                            f"{qid}: parsed={parsed:.4f} but Pairwise_Values={actual:.4f}"
                        )
                elif parsed is None:
                    parse_issues.append(f"{qid}: could not parse text '{str(text_val)[:60]}'")
                errors.extend(parse_errs)

    if parse_issues:
        notes.append(f"Text parse mismatches: {'; '.join(parse_issues)}")

    # --- 7. Build 5x5 matrix ---
    criteria_order = ["C1", "C2", "C3", "C4", "C5"]
    matrix = build_pairwise_matrix(pairwise_values, criteria_order)

    # --- 8. Compute consistency ---
    weights = calculate_ahp_weights(matrix)
    consistency = calculate_consistency(matrix, weights)

    # Cross-check CR with AHP_Validation sheet
    if valid_row is not None:
        cr_from_sheet = valid_row.get("CR")
        if cr_from_sheet is not None:
            cr_diff = abs(consistency.cr - float(cr_from_sheet))
            if cr_diff > 1e-6:
                notes.append(
                    f"CR re-calculated={consistency.cr:.6f} differs from "
                    f"AHP_Validation CR={float(cr_from_sheet):.6f} (Δ={cr_diff:.1e})"
                )

    # Check for empty values in matrix
    empty_pairs = [qid for qid, val in pairwise_values.items() if val is None]
    if empty_pairs:
        errors.append(f"Missing values for: {', '.join(empty_pairs)}")

    pw = PairwiseValues(
        respondent_id=respondent_id,
        original_code=original_code,
        source_type=source_type,
        values=pairwise_values,
        matrix=matrix,
    )

    return RespondentJudgement(
        profile=profile,
        pairwise=pw,
        consistency=consistency,
        raw_text_answers=raw_text_answers,
        parse_errors=errors,
        validation_notes=notes,
    )


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def to_serializable(obj):
    """Convert dataclass/non-serializable to dict/primitive."""
    if isinstance(obj, (datetime,)):
        return obj.isoformat()
    if hasattr(obj, "__dataclass_fields__"):
        return {k: to_serializable(v) for k, v in asdict(obj).items()}
    if isinstance(obj, list):
        return [to_serializable(v) for v in obj]
    if isinstance(obj, dict):
        return {k: to_serializable(v) for k, v in obj.items()}
    return obj


# ---------------------------------------------------------------------------
# Output generation
# ---------------------------------------------------------------------------


def write_responses_csv(
    path: Path,
    respondents: list[RespondentJudgement],
) -> Path:
    """Write expert_judgement_responses.csv (all respondents)."""
    fieldnames = [
        "respondent_id", "original_code", "source_type",
        "role_category", "expertise", "education", "experience",
        "spotify_status", "spotify_frequency",
        "criteria_adequacy", "top_criterion", "top_reason",
        "cr", "is_consistent",
        "c1_weight", "c2_weight", "c3_weight", "c4_weight", "c5_weight",
        "top_criterion_weight",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in respondents:
            weights = calculate_ahp_weights(r.pairwise.matrix)
            writer.writerow({
                "respondent_id": r.profile.respondent_id,
                "original_code": r.profile.original_code,
                "source_type": r.profile.source_type,
                "role_category": r.profile.role_category,
                "expertise": r.profile.expertise,
                "education": r.profile.education,
                "experience": r.profile.experience,
                "spotify_status": r.profile.spotify_status,
                "spotify_frequency": r.profile.spotify_frequency,
                "criteria_adequacy": r.profile.criteria_adequacy,
                "top_criterion": r.profile.top_criterion,
                "top_reason": r.profile.top_reason,
                "cr": r.consistency.cr,
                "is_consistent": r.consistency.is_consistent,
                "c1_weight": weights[0],
                "c2_weight": weights[1],
                "c3_weight": weights[2],
                "c4_weight": weights[3],
                "c5_weight": weights[4],
                "top_criterion_weight": max(weights),
            })
    return path


def write_valid_responses_csv(
    path: Path,
    respondents: list[RespondentJudgement],
) -> Path:
    """Write expert_judgement_valid_responses.csv (CR <= 0.10 only)."""
    valid = [r for r in respondents if r.consistency.is_consistent]
    return write_responses_csv(path, valid)


def write_pairwise_matrices_json(
    path: Path,
    respondents: list[RespondentJudgement],
) -> Path:
    """Write expert_judgement_pairwise_matrices.json."""
    data = []
    for r in respondents:
        weights = calculate_ahp_weights(r.pairwise.matrix)
        data.append({
            "respondent_id": r.profile.respondent_id,
            "original_code": r.profile.original_code,
            "source_type": r.profile.source_type,
            "pairwise_values": {
                qid: r.pairwise.values.get(qid)
                for qid, _, _, _ in PAIRWISE_QUESTIONS
            },
            "pairwise_matrix": r.pairwise.matrix,
            "criteria": [{"id": k, "name": v} for k, v in CRITERIA.items()],
            "weights": [
                {"criterion_id": cid, "criterion_name": CRITERIA[cid], "weight": weights[i]}
                for i, cid in enumerate(["C1", "C2", "C3", "C4", "C5"])
            ],
        })
    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_respondents": len(data),
        "criteria_count": len(CRITERIA),
        "pairwise_questions": [
            {"id": qid, "criterion_a": CRITERIA[c_a], "criterion_b": CRITERIA[c_b]}
            for qid, c_a, c_b, _ in PAIRWISE_QUESTIONS
        ],
        "respondents": data,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    return path


def write_validation_summary_json(
    path: Path,
    respondents: list[RespondentJudgement],
) -> Path:
    """Write expert_judgement_validation_summary.json."""
    valid_count = sum(1 for r in respondents if r.consistency.is_consistent)
    invalid_count = sum(1 for r in respondents if not r.consistency.is_consistent)
    actual_count = sum(1 for r in respondents if r.profile.source_type == "actual")
    synthetic_count = sum(1 for r in respondents if r.profile.source_type == "synthetic")

    data = []
    for r in respondents:
        data.append({
            "respondent_id": r.profile.respondent_id,
            "original_code": r.profile.original_code,
            "source_type": r.profile.source_type,
            "cr": r.consistency.cr,
            "lambda_max": r.consistency.lambda_max,
            "ci": r.consistency.ci,
            "is_consistent": r.consistency.is_consistent,
            "consistency_threshold": CR_THRESHOLD,
            "validation_notes": r.validation_notes,
            "error_count": len(r.parse_errors),
            "parse_errors": r.parse_errors,
        })

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "validation_method": "AHP Consistency Ratio (Saaty)",
        "cr_threshold": CR_THRESHOLD,
        "criteria_count": len(CRITERIA),
        "total_respondents": len(respondents),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "actual_count": actual_count,
        "synthetic_count": synthetic_count,
        "respondents": data,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    return path


def write_mapping_summary_json(
    path: Path,
    respondents: list[RespondentJudgement],
    spreadsheet_path: str,
) -> Path:
    """Write expert_judgement_mapping_summary.json."""
    valid = [r for r in respondents if r.consistency.is_consistent]
    invalid = [r for r in respondents if not r.consistency.is_consistent]
    actual = [r for r in respondents if r.profile.source_type == "actual"]
    synthetic = [r for r in respondents if r.profile.source_type == "synthetic"]

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "task": "MS-17A — Expert Judgement Spreadsheet Intake & Mapping",
        "spreadsheet_path": spreadsheet_path,
        "criteria_mapping": {k: v for k, v in CRITERIA.items()},
        "pairwise_mapping": [
            {
                "id": qid,
                "pair": f"{CRITERIA[c_a]} vs {CRITERIA[c_b]}",
                "column_in_pairwise_sheet": col,
            }
            for qid, c_a, c_b, col in PAIRWISE_QUESTIONS
        ],
        "respondent_summary": {
            "total": len(respondents),
            "actual": len(actual),
            "synthetic": len(synthetic),
            "valid": len(valid),
            "invalid": len(invalid),
        },
        "invalid_respondents": [
            {
                "respondent_id": r.profile.respondent_id,
                "original_code": r.profile.original_code,
                "source_type": r.profile.source_type,
                "cr": r.consistency.cr,
                "reason": "CR exceeds threshold",
                "notes": r.validation_notes,
            }
            for r in invalid
        ],
        "data_status": {
            "actual_used": len(actual),
            "synthetic_used": len(synthetic),
            "note": "Synthetic data must not be reported as real expert judgement.",
        },
        "next_step": "MS-17B — AHP and Fuzzy AHP Aggregation & Ranking",
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    return path


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="MS-17A: Expert Judgement Spreadsheet Intake & Mapping",
    )
    parser.add_argument(
        "--spreadsheet",
        default=str(
            PROJECT_ROOT / "datasets" / "external" / "expert_judgement"
            / "spotify_expert_judgement_augmented.xlsx"
        ),
        help="Path to augmented Google Form spreadsheet",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "datasets" / "processed" / "expert_judgement"),
        help="Output directory for processed files",
    )
    args = parser.parse_args()

    spreadsheet_path = Path(args.spreadsheet).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"📄 Loading spreadsheet: {spreadsheet_path}")
    wb = load_spreadsheet(spreadsheet_path)

    print("📋 Reading Form Responses 2...")
    form_responses = load_form_responses(wb["Form Responses 2"])
    print(f"   → {len(form_responses)} response rows found")

    print("📋 Reading Pairwise_Values...")
    pairwise_data = load_pairwise_values(wb["Pairwise_Values"])
    print(f"   → {len(pairwise_data)} rows found")

    print("📋 Reading AHP_Validation...")
    validation_data = load_ahp_validation(wb["AHP_Validation"])
    print(f"   → {len(validation_data)} rows found")

    # Process each respondent
    respondents: list[RespondentJudgement] = []
    for i, (row_num, form_row) in enumerate(sorted(form_responses.items()), start=1):
        pair_row = pairwise_data.get(row_num)
        valid_row = validation_data.get(row_num)

        judgement = process_respondent(row_num, form_row, pair_row, valid_row, i)
        respondents.append(judgement)

        # Print summary for each
        src = judgement.profile.source_type.upper()
        cons = "✅" if judgement.consistency.is_consistent else "❌"
        cr = judgement.consistency.cr
        name_display = (form_row.get("Nama Responden  ") or "?").strip()[:20]
        print(
            f"  {judgement.profile.respondent_id} "
            f"[{src}] {name_display:<22s} "
            f"CR={cr:.4f} {cons}"
        )

    # --- Generate outputs ---
    print(f"\n📝 Writing outputs to {output_dir}")

    p1 = write_responses_csv(output_dir / "expert_judgement_responses.csv", respondents)
    print(f"  ✅ {p1.name}")

    p2 = write_valid_responses_csv(output_dir / "expert_judgement_valid_responses.csv", respondents)
    print(f"  ✅ {p2.name}")

    p3 = write_pairwise_matrices_json(output_dir / "expert_judgement_pairwise_matrices.json", respondents)
    print(f"  ✅ {p3.name}")

    p4 = write_validation_summary_json(output_dir / "expert_judgement_validation_summary.json", respondents)
    print(f"  ✅ {p4.name}")

    p5 = write_mapping_summary_json(output_dir / "expert_judgement_mapping_summary.json", respondents, str(spreadsheet_path))
    print(f"  ✅ {p5.name}")

    # --- Summary ---
    valid_count = sum(1 for r in respondents if r.consistency.is_consistent)
    invalid_count = sum(1 for r in respondents if not r.consistency.is_consistent)
    actual_count = sum(1 for r in respondents if r.profile.source_type == "actual")
    synthetic_count = sum(1 for r in respondents if r.profile.source_type == "synthetic")

    print(f"\n{'='*50}")
    print(f"📊 SUMMARY")
    print(f"  Total respondents : {len(respondents)}")
    print(f"  Actual            : {actual_count}")
    print(f"  Synthetic         : {synthetic_count}")
    print(f"  Valid (CR≤0.10)   : {valid_count}")
    print(f"  Invalid (CR>0.10) : {invalid_count}")
    print(f"{'='*50}")

    # Warn about synthetic data
    if synthetic_count > 0:
        print("\n⚠️  NOTE: Synthetic respondents are present. Do NOT report them as real")
        print("   expert judgement in thesis or publications.")

    print("\n✅ MS-17A complete. Ready for MS-17B (AHP/Fuzzy AHP aggregation).")


if __name__ == "__main__":
    main()
